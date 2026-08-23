# -*- coding: utf-8 -*-
"""
Sincronizador Automático de Leads y Traza en Vivo hacia Registro_General_ProtegeDatosLocal.xlsx
Uso: python sincronizar_leads.py
"""
import subprocess
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("========================================================================")
print("     CONSULTANDO REGISTROS EN VIVO DESDE HOSTINGER (PROTEGEDATOSLOCAL)   ")
print("========================================================================")

leads = []
try:
    cmd = 'powershell -Command "ssh hostinger \"cat ~/pdl_secure_data/traza_accesos.log domains/inncivica.cloud/public_html/protegedatoslocal/api/traza_segura.php 2>/dev/null\""'
    res = subprocess.run(cmd, shell=True, capture_output=True, text=True, encoding="utf-8")
    
    seen = set()
    for line in res.stdout.strip().splitlines():
        if not line.strip() or line.startswith("<?php"):
            continue
        try:
            data = json.loads(line)
            key = f"{data.get('fecha')}_{data.get('email')}"
            if key not in seen:
                seen.add(key)
                leads.append(data)
        except Exception:
            pass
    print(f"-> Conectado a Hostinger. Total de funcionarios registrados: {len(leads)}\n")
except Exception as e:
    print(f"[!] Error de conexion: {e}")

if leads:
    print(f"{'FECHA':<20} | {'MUNICIPALIDAD':<32} | {'FUNCIONARIO (CARGO)':<35} | {'CORREO'}")
    print("-" * 110)
    for lead in leads:
        fecha = lead.get('fecha', '-')
        muni = lead.get('municipio', '-')
        nom = f"{lead.get('nombre', '-')} ({lead.get('cargo', '-')})"
        email = lead.get('email', '-')
        print(f"{fecha:<20} | {muni:<32} | {nom:<35} | {email}")
else:
    print("Aún no hay registros en el servidor remoto.")

# Actualizar el Excel Maestro
excel_file = Path("Registro_General_ProtegeDatosLocal.xlsx")
if excel_file.exists():
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws1 = wb["Leads y Registros en Vivo"] if "Leads y Registros en Vivo" in wb.sheetnames else wb.active
        
        while ws1.max_row > 1:
            ws1.delete_rows(2)
            
        font_cell = Font(name='Calibri', size=10, color='0F172A')
        font_bold = Font(name='Calibri', size=10, bold=True, color='0A2540')
        fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        border_thin = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        for idx, lead in enumerate(leads, start=1):
            row_data = [
                idx,
                lead.get('fecha', '-'),
                lead.get('municipio', '-'),
                lead.get('nombre', '-'),
                lead.get('cargo', '-'),
                lead.get('email', '-'),
                lead.get('departamento', 'General'),
                "Nuevo / Diagnóstico Iniciado",
                "Remitir Informe Ejecutivo y TDR en Word",
                "Honorarios Suma Alzada (Art. 4 Ley 18.883)"
            ]
            ws1.append(row_data)
            
        for row_idx in range(2, ws1.max_row + 1):
            is_zebra = (row_idx % 2 == 0)
            for col_idx in range(1, 11):
                cell = ws1.cell(row=row_idx, column=col_idx)
                cell.font = font_bold if col_idx in [3, 4] else font_cell
                cell.border = border_thin
                if is_zebra:
                    cell.fill = fill_zebra
                    
        wb.save(excel_file)
        print("\n[OK] Excel Maestro 'Registro_General_ProtegeDatosLocal.xlsx' actualizado con éxito!")
    except Exception as e:
        print(f"\n[!] No se pudo actualizar el Excel: {e}")

print("========================================================================")
